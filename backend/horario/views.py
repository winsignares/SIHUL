from django.shortcuts import render
from .models import Horario, HorarioFusionado, HorarioEstudiante, SolicitudEspacio
from grupos.models import Grupo
from asignaturas.models import Asignatura
from usuarios.models import Usuario
from espacios.models import EspacioFisico
from django.http import JsonResponse, HttpResponse
import json
from django.views.decorators.csrf import csrf_exempt
import datetime
from notificaciones.signals import crear_notificacion
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import sys
import openpyxl

# ---------- Endpoint para Mi Horario (Docente) ----------
@csrf_exempt
def mi_horario_docente(request):
    """Retorna el horario del docente logueado"""
    if request.method != 'GET':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        # Obtener usuario_id de headers o query params
        usuario_id = request.GET.get('usuario_id') or request.headers.get('X-Usuario-Id')
        
        if not usuario_id:
            return JsonResponse({"error": "usuario_id es requerido"}, status=400)
        
        # Verificar que el usuario existe
        try:
            docente = Usuario.objects.get(id=usuario_id)
        except Usuario.DoesNotExist:
            return JsonResponse({"error": "Usuario no encontrado"}, status=404)
        
        # Obtener horarios del docente con información extendida
        horarios = Horario.objects.filter(
            docente=docente,
            estado='aprobado'
        ).select_related('grupo', 'asignatura', 'espacio', 'grupo__programa').all()
        
        lst = []
        for h in horarios:
            lst.append({
                "id": h.id,
                "diaSemana": h.dia_semana,
                "horaInicio": str(h.hora_inicio),
                "horaFin": str(h.hora_fin),
                "asignaturaId": h.asignatura.id,
                "asignatura": h.asignatura.nombre,
                "grupoId": h.grupo.id,
                "grupo": h.grupo.nombre,
                "espacioId": h.espacio.id,
                "espacio": h.espacio.nombre,
                "docenteId": h.docente.id if h.docente else None,
                "docente": h.docente.nombre if h.docente else "Sin asignar",
                "cantidadEstudiantes": h.cantidad_estudiantes,
                "programa": h.grupo.programa.nombre if h.grupo.programa else None,
                "semestre": h.grupo.semestre
            })
        
        return JsonResponse({"horarios": lst}, status=200)
        
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

# ---------- Endpoint para Mi Horario (Estudiante) ----------
@csrf_exempt
def mi_horario_estudiante(request):
    """Retorna el horario del estudiante logueado"""
    if request.method != 'GET':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        # Obtener usuario_id de headers o query params
        usuario_id = request.GET.get('usuario_id') or request.headers.get('X-Usuario-Id')
        
        if not usuario_id:
            return JsonResponse({"error": "usuario_id es requerido"}, status=400)
        
        # Verificar que el usuario existe
        try:
            estudiante = Usuario.objects.get(id=usuario_id)
        except Usuario.DoesNotExist:
            return JsonResponse({"error": "Usuario no encontrado"}, status=404)
        
        # Obtener horarios del estudiante
        inscripciones = HorarioEstudiante.objects.filter(estudiante=estudiante).select_related('horario', 'horario__grupo', 'horario__asignatura', 'horario__espacio', 'horario__docente', 'horario__grupo__programa')
        
        lst = []
        for insc in inscripciones:
            h = insc.horario
            lst.append({
                "id": h.id,
                "diaSemana": h.dia_semana,
                "horaInicio": str(h.hora_inicio),
                "horaFin": str(h.hora_fin),
                "asignaturaId": h.asignatura.id,
                "asignatura": h.asignatura.nombre,
                "grupoId": h.grupo.id,
                "grupo": h.grupo.nombre,
                "espacioId": h.espacio.id,
                "espacio": h.espacio.nombre,
                "docenteId": h.docente.id if h.docente else None,
                "docente": h.docente.nombre if h.docente else "Sin asignar",
                "cantidadEstudiantes": h.cantidad_estudiantes,
                "programa": h.grupo.programa.nombre if h.grupo.programa else None,
                "semestre": h.grupo.semestre
            })
        
        return JsonResponse({"horarios": lst}, status=200)
        
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def inscribir_estudiante(request):
    """Inscribe un estudiante a un horario"""
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        data = json.loads(request.body)
        usuario_id = data.get('usuario_id')
        horario_id = data.get('horario_id')
        
        if not usuario_id or not horario_id:
            return JsonResponse({"error": "usuario_id y horario_id son requeridos"}, status=400)
            
        estudiante = Usuario.objects.get(id=usuario_id)
        horario = Horario.objects.get(id=horario_id)
        
        # Crear inscripcion
        inscripcion, created = HorarioEstudiante.objects.get_or_create(estudiante=estudiante, horario=horario)
        
        if created:
            return JsonResponse({"message": "Estudiante inscrito correctamente"}, status=201)
        else:
            return JsonResponse({"message": "El estudiante ya estaba inscrito"}, status=200)
            
    except Usuario.DoesNotExist:
        return JsonResponse({"error": "Estudiante no encontrado"}, status=404)
    except Horario.DoesNotExist:
        return JsonResponse({"error": "Horario no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

# ---------- Horario CRUD ----------
@csrf_exempt
def create_horario(request):
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body)
        grupo_id = data.get('grupo_id')
        asignatura_id = data.get('asignatura_id')
        espacio_id = data.get('espacio_id')
        dia_semana = data.get('dia_semana')
        hora_inicio = data.get('hora_inicio')
        hora_fin = data.get('hora_fin')
        docente_id = data.get('docente_id')
        cantidad = data.get('cantidad_estudiantes')
        usuario_id = data.get('usuario_id')  # Usuario que hace la solicitud
        
        if not grupo_id or not asignatura_id or not espacio_id or not dia_semana or not hora_inicio or not hora_fin:
            return JsonResponse({"error": "Faltan campos requeridos"}, status=400)
        
        grupo = Grupo.objects.get(id=grupo_id)
        asignatura = Asignatura.objects.get(id=asignatura_id)
        espacio = EspacioFisico.objects.get(id=espacio_id)
        docente = Usuario.objects.get(id=docente_id) if docente_id else None
        usuario = Usuario.objects.get(id=usuario_id) if usuario_id else None
        
        hi = datetime.time.fromisoformat(hora_inicio)
        hf = datetime.time.fromisoformat(hora_fin)
        
        # Verificar si el usuario es planificador
        es_planificador = usuario and usuario.rol and usuario.rol.nombre == 'planeacion_facultad'
        
        if es_planificador:
            # Crear solicitud de espacio en lugar de horario directo
            solicitud = SolicitudEspacio(
                grupo=grupo,
                asignatura=asignatura,
                docente=docente,
                espacio_solicitado=espacio,
                planificador=usuario,
                dia_semana=dia_semana,
                hora_inicio=hi,
                hora_fin=hf,
                cantidad_estudiantes=int(cantidad) if cantidad is not None else None,
                estado='pendiente'
            )
            solicitud.save()
            
            # Crear notificación para administradores
            admins = Usuario.objects.filter(rol__nombre='admin')
            for admin in admins:
                crear_notificacion(
                    id_usuario=admin.id,
                    tipo='solicitud_espacio',
                    mensaje=f'Nueva solicitud de espacio (ID: {solicitud.id}): {asignatura.nombre} - Grupo {grupo.nombre} - Aula: {espacio.nombre} - {dia_semana} {hi}-{hf}',
                    prioridad='alta'
                )
            
            return JsonResponse({
                "message": "Solicitud de espacio creada exitosamente",
                "id": solicitud.id,
                "tipo": "solicitud"
            }, status=201)
        else:
            # Admin: crear horario directamente
            h = Horario(
                grupo=grupo,
                asignatura=asignatura,
                docente=docente,
                espacio=espacio,
                dia_semana=dia_semana,
                hora_inicio=hi,
                hora_fin=hf,
                cantidad_estudiantes=int(cantidad) if cantidad is not None else None,
                estado='aprobado'
            )
            h.save()
            return JsonResponse({
                "message": "Horario creado",
                "id": h.id,
                "tipo": "horario"
            }, status=201)
    except (Grupo.DoesNotExist, Asignatura.DoesNotExist, EspacioFisico.DoesNotExist, Usuario.DoesNotExist):
        return JsonResponse({"error": "Relacionada no encontrada."}, status=404)
    except ValueError:
        return JsonResponse({"error": "Formato de hora inválido o valor numérico incorrecto."}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def update_horario(request):
    if request.method != 'PUT':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body)
        id = data.get('id')
        if not id:
            return JsonResponse({"error": "ID es requerido"}, status=400)
        h = Horario.objects.get(id=id)
        if 'grupo_id' in data:
            h.grupo = Grupo.objects.get(id=data.get('grupo_id'))
        if 'asignatura_id' in data:
            h.asignatura = Asignatura.objects.get(id=data.get('asignatura_id'))
        if 'docente_id' in data:
            h.docente = Usuario.objects.get(id=data.get('docente_id')) if data.get('docente_id') else None
        if 'espacio_id' in data:
            h.espacio = EspacioFisico.objects.get(id=data.get('espacio_id'))
        if 'dia_semana' in data:
            h.dia_semana = data.get('dia_semana')
        if 'hora_inicio' in data:
            h.hora_inicio = datetime.time.fromisoformat(data.get('hora_inicio'))
        if 'hora_fin' in data:
            h.hora_fin = datetime.time.fromisoformat(data.get('hora_fin'))
        if 'cantidad_estudiantes' in data:
            h.cantidad_estudiantes = int(data.get('cantidad_estudiantes')) if data.get('cantidad_estudiantes') is not None else None
        h.save()
        return JsonResponse({"message": "Horario actualizado", "id": h.id}, status=200)
    except Horario.DoesNotExist:
        return JsonResponse({"error": "Horario no encontrado."}, status=404)
    except (Grupo.DoesNotExist, Asignatura.DoesNotExist, EspacioFisico.DoesNotExist, Usuario.DoesNotExist):
        return JsonResponse({"error": "Relacionada no encontrada."}, status=404)
    except ValueError:
        return JsonResponse({"error": "Formato de hora inválido o valor numérico incorrecto."}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def delete_horario(request):
    if request.method != 'DELETE':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body)
        id = data.get('id')
        if not id:
            return JsonResponse({"error": "ID es requerido"}, status=400)
        h = Horario.objects.get(id=id)
        h.delete()
        return JsonResponse({"message": "Horario eliminado"}, status=200)
    except Horario.DoesNotExist:
        return JsonResponse({"error": "Horario no encontrado."}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
 
@csrf_exempt
def get_horario(request, id=None):
    if id is None:
        return JsonResponse({"error": "El ID es requerido en la URL"}, status=400)
    try:
        h = Horario.objects.get(id=id)
        return JsonResponse({
            "id": h.id,
            "grupo_id": h.grupo.id,
            "asignatura_id": h.asignatura.id,
            "docente_id": (h.docente.id if h.docente else None),
            "espacio_id": h.espacio.id,
            "dia_semana": h.dia_semana,
            "hora_inicio": str(h.hora_inicio),
            "hora_fin": str(h.hora_fin),
            "cantidad_estudiantes": h.cantidad_estudiantes
        }, status=200)
    except Horario.DoesNotExist:
        return JsonResponse({"error": "Horario no encontrado."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def list_horarios(request):
    if request.method == 'GET':
        items = Horario.objects.filter(estado='aprobado')
        lst = [{
            "id": i.id,
            "grupo_id": i.grupo.id,
            "asignatura_id": i.asignatura.id,
            "docente_id": (i.docente.id if i.docente else None),
            "espacio_id": i.espacio.id,
            "dia_semana": i.dia_semana,
            "hora_inicio": str(i.hora_inicio),
            "hora_fin": str(i.hora_fin),
            "cantidad_estudiantes": i.cantidad_estudiantes
        } for i in items]
        return JsonResponse({"horarios": lst}, status=200)

@csrf_exempt
def list_horarios_extendidos(request):
    """Lista horarios con información extendida (nombres de relaciones) - Solo aprobados"""
    if request.method == 'GET':
        # Traer solo horarios aprobados
        items = Horario.objects.select_related('grupo', 'asignatura', 'docente', 'espacio', 'grupo__programa').filter(estado='aprobado')
        lst = []
        for i in items:
            lst.append({
                "id": i.id,
                "grupo_id": i.grupo.id,
                "grupo_nombre": i.grupo.nombre,
                "programa_id": i.grupo.programa.id,
                "programa_nombre": i.grupo.programa.nombre,
                "semestre": i.grupo.semestre,
                "asignatura_id": i.asignatura.id,
                "asignatura_nombre": i.asignatura.nombre,
                "docente_id": (i.docente.id if i.docente else None),
                "docente_nombre": i.docente.nombre if i.docente else "Sin asignar",
                "espacio_id": i.espacio.id,
                "espacio_nombre": i.espacio.nombre,
                "dia_semana": i.dia_semana,
                "hora_inicio": str(i.hora_inicio),
                "hora_fin": str(i.hora_fin),
                "cantidad_estudiantes": i.cantidad_estudiantes,
                "estado": i.estado
            })
        return JsonResponse({"horarios": lst}, status=200)

# ---------- HorarioFusionado CRUD ----------
# ---------- Horario_Fusionado CRUD ----------
@csrf_exempt
def create_horario_fusionado(request):
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body)
        grupo1_id = data.get('grupo1_id')
        grupo2_id = data.get('grupo2_id')
        grupo3_id = data.get('grupo3_id')
        asignatura_id = data.get('asignatura_id')
        espacio_id = data.get('espacio_id')
        dia_semana = data.get('dia_semana')
        hora_inicio = data.get('hora_inicio')
        hora_fin = data.get('hora_fin')
        docente_id = data.get('docente_id')
        cantidad = data.get('cantidad_estudiantes')
        comentario = data.get('comentario')
        if not grupo1_id or not grupo2_id or not asignatura_id or not espacio_id or not dia_semana or not hora_inicio or not hora_fin:
            return JsonResponse({"error": "Faltan campos requeridos"}, status=400)
        grupo1 = Grupo.objects.get(id=grupo1_id)
        grupo2 = Grupo.objects.get(id=grupo2_id)
        grupo3 = Grupo.objects.get(id=grupo3_id) if grupo3_id else None
        asignatura = Asignatura.objects.get(id=asignatura_id)
        espacio = EspacioFisico.objects.get(id=espacio_id)
        docente = Usuario.objects.get(id=docente_id) if docente_id else None
        hi = datetime.time.fromisoformat(hora_inicio)
        hf = datetime.time.fromisoformat(hora_fin)
        hfus = HorarioFusionado(grupo1=grupo1, grupo2=grupo2, grupo3=grupo3, asignatura=asignatura, docente=docente, espacio=espacio, dia_semana=dia_semana, hora_inicio=hi, hora_fin=hf, cantidad_estudiantes=(int(cantidad) if cantidad is not None else None), comentario=comentario)
        hfus.save()
        return JsonResponse({"message": "Horario fusionado creado", "id": hfus.id}, status=201)
    except (Grupo.DoesNotExist, Asignatura.DoesNotExist, EspacioFisico.DoesNotExist, Usuario.DoesNotExist):
        return JsonResponse({"error": "Relacionada no encontrada."}, status=404)
    except ValueError:
        return JsonResponse({"error": "Formato de hora inválido o valor numérico incorrecto."}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def update_horario_fusionado(request):
    if request.method != 'PUT':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body)
        id = data.get('id')
        if not id:
            return JsonResponse({"error": "ID es requerido"}, status=400)
        h = HorarioFusionado.objects.get(id=id)
        if 'grupo1_id' in data:
            h.grupo1 = Grupo.objects.get(id=data.get('grupo1_id'))
        if 'grupo2_id' in data:
            h.grupo2 = Grupo.objects.get(id=data.get('grupo2_id'))
        if 'grupo3_id' in data:
            h.grupo3 = Grupo.objects.get(id=data.get('grupo3_id')) if data.get('grupo3_id') else None
        if 'asignatura_id' in data:
            h.asignatura = Asignatura.objects.get(id=data.get('asignatura_id'))
        if 'docente_id' in data:
            h.docente = Usuario.objects.get(id=data.get('docente_id')) if data.get('docente_id') else None
        if 'espacio_id' in data:
            h.espacio = EspacioFisico.objects.get(id=data.get('espacio_id'))
        if 'dia_semana' in data:
            h.dia_semana = data.get('dia_semana')
        if 'hora_inicio' in data:
            h.hora_inicio = datetime.time.fromisoformat(data.get('hora_inicio'))
        if 'hora_fin' in data:
            h.hora_fin = datetime.time.fromisoformat(data.get('hora_fin'))
        if 'cantidad_estudiantes' in data:
            h.cantidad_estudiantes = int(data.get('cantidad_estudiantes')) if data.get('cantidad_estudiantes') is not None else None
        if 'comentario' in data:
            h.comentario = data.get('comentario')
        h.save()
        return JsonResponse({"message": "Horario fusionado actualizado", "id": h.id}, status=200)
    except HorarioFusionado.DoesNotExist:
        return JsonResponse({"error": "Horario fusionado no encontrado."}, status=404)
    except (Grupo.DoesNotExist, Asignatura.DoesNotExist, EspacioFisico.DoesNotExist, Usuario.DoesNotExist):
        return JsonResponse({"error": "Relacionada no encontrada."}, status=404)
    except ValueError:
        return JsonResponse({"error": "Formato de hora inválido o valor numérico incorrecto."}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def delete_horario_fusionado(request):
    if request.method != 'DELETE':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body)
        id = data.get('id')
        if not id:
            return JsonResponse({"error": "ID es requerido"}, status=400)
        h = HorarioFusionado.objects.get(id=id)
        h.delete()
        return JsonResponse({"message": "Horario fusionado eliminado"}, status=200)
    except HorarioFusionado.DoesNotExist:
        return JsonResponse({"error": "Horario fusionado no encontrado."}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    
@csrf_exempt
def get_horario_fusionado(request, id=None):
    if id is None:
        return JsonResponse({"error": "El ID es requerido en la URL"}, status=400)
    try:
        h = HorarioFusionado.objects.get(id=id)
        return JsonResponse({
            "id": h.id,
            "grupo1_id": h.grupo1.id,
            "grupo2_id": h.grupo2.id,
            "grupo3_id": (h.grupo3.id if h.grupo3 else None),
            "asignatura_id": h.asignatura.id,
            "docente_id": (h.docente.id if h.docente else None),
            "espacio_id": h.espacio.id,
            "dia_semana": h.dia_semana,
            "hora_inicio": str(h.hora_inicio),
            "hora_fin": str(h.hora_fin),
            "cantidad_estudiantes": h.cantidad_estudiantes,
            "comentario": h.comentario
        }, status=200)
    except HorarioFusionado.DoesNotExist:
        return JsonResponse({"error": "Horario fusionado no encontrado."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def list_horarios_fusionados(request):
    if request.method == 'GET':
        items = HorarioFusionado.objects.all()
        lst = [{
            "id": i.id,
            "grupo1_id": i.grupo1.id,
            "grupo2_id": i.grupo2.id,
            "grupo3_id": (i.grupo3.id if i.grupo3 else None),
            "asignatura_id": i.asignatura.id,
            "docente_id": (i.docente.id if i.docente else None),
            "espacio_id": i.espacio.id,
            "dia_semana": i.dia_semana,
            "hora_inicio": str(i.hora_inicio),
            "hora_fin": str(i.hora_fin),
            "cantidad_estudiantes": i.cantidad_estudiantes,
            "comentario": i.comentario
        } for i in items]
        return JsonResponse({"horarios_fusionados": lst}, status=200)

# -------- Funciones auxiliares para exportación --------
def obtener_horarios_usuario(usuario_id):
    """
    Obtiene los horarios de un usuario, ya sea docente o estudiante
    """
    try:
        usuario = Usuario.objects.get(id=usuario_id)
        rol_nombre = usuario.rol.nombre if usuario.rol else None
        
        if rol_nombre == 'estudiante':
            # Obtener horarios del estudiante a través de HorarioEstudiante
            inscripciones = HorarioEstudiante.objects.filter(
                estudiante=usuario
            ).select_related(
                'horario__grupo',
                'horario__asignatura', 
                'horario__espacio',
                'horario__docente'
            )
            horarios = [insc.horario for insc in inscripciones]
            return horarios, rol_nombre
        else:
            # Asumir que es docente
            horarios = Horario.objects.filter(docente=usuario, estado='aprobado').select_related(
                'grupo', 'asignatura', 'espacio', 'docente'
            )
            return list(horarios), rol_nombre
    except Usuario.DoesNotExist:
        return [], None



def generar_tabla_horario(horarios, rol_nombre):
    """
    Organiza los horarios en una estructura de tabla (días x horas)
    """
    # Días de la semana
    dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
    
    # Obtener rango de horas (de 7:00 a 20:00 por defecto)
    horas_inicio = set()
    horas_fin = set()
    
    for h in horarios:
        horas_inicio.add(h.hora_inicio.hour)
        horas_fin.add(h.hora_fin.hour)
    
    hora_min = min(horas_inicio) if horas_inicio else 7
    hora_max = max(horas_fin) if horas_fin else 20
    
    # Crear franjas horarias cada 1 hora
    franjas = []
    for h in range(hora_min, hora_max):
        franjas.append(f"{h:02d}:00 - {h+1:02d}:00")
    
    # Crear diccionario de celdas (dia, franja) -> info
    tabla = {}
    dia_map = {
        'Lunes': 0, 'Martes': 1, 'Miércoles': 2,
        'Jueves': 3, 'Viernes': 4, 'Sábado': 5
    }
    
    for h in horarios:
        dia_idx = dia_map.get(h.dia_semana)
        if dia_idx is None:
            continue
        
        # Encontrar la franja correspondiente
        hora_clase = h.hora_inicio.hour
        franja_idx = hora_clase - hora_min
        
        if 0 <= franja_idx < len(franjas):
            key = (dia_idx, franja_idx)
            info = f"{h.asignatura.nombre}\n{h.grupo.nombre}\n{h.espacio.nombre}"
            if rol_nombre == 'estudiante' and h.docente:
                info += f"\n{h.docente.nombre}"
            tabla[key] = info
    
    return dias, franjas, tabla


# -------- Endpoint para exportar PDF --------
@csrf_exempt
def exportar_horario_pdf(request):
    """Genera y descarga el horario en formato PDF"""
    if request.method != 'GET':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    usuario_id = request.GET.get('usuario_id')
    if not usuario_id:
        return JsonResponse({"error": "usuario_id es requerido"}, status=400)
    
    horarios, rol_nombre = obtener_horarios_usuario(usuario_id)
    
    if not horarios:
        return JsonResponse({"error": "No se encontraron horarios"}, status=404)
    
    # Generar tabla
    dias, franjas, tabla = generar_tabla_horario(horarios, rol_nombre)
    
    # Crear PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           rightMargin=30, leftMargin=30,
                           topMargin=30, bottomMargin=30)
    
    # Preparar datos para la tabla
    data = [['Hora'] + dias]  # Header
    
    for franja_idx, franja in enumerate(franjas):
        row = [franja]
        for dia_idx in range(len(dias)):
            key = (dia_idx, franja_idx)
            cell_text = tabla.get(key, '')
            row.append(cell_text)
        data.append(row)
    
    # Crear tabla con anchos uniformes
    # Calcular ancho disponible (landscape A4 = 11.69 inches - margins)
    available_width = 11.69*inch - 60  # 60 points total margins
    col_width = available_width / 7  # 7 columnas (1 hora + 6 días)
    
    table = Table(data, colWidths=[col_width]*7, rowHeights=[0.7*inch]*(len(data)))
    
    # Estilo de tabla
    style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#991B1B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Columna de horas
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F3F4F6')),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (0, -1), 9),
        
        # Celdas de datos
        ('FONTSIZE', (1, 1), (-1, -1), 7),
        ('TOPPADDING', (1, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (1, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
    ])

    
    # Color para celdas con clases
    for (dia_idx, franja_idx), info in tabla.items():
        if info:
            col = dia_idx + 1  # +1 porque la primera columna es de horas
            row = franja_idx + 1  # +1 porque la primera fila es header
            style.add('BACKGROUND', (col, row), (col, row), colors.HexColor('#DBEAFE'))
    
    table.setStyle(style)
    
    # Generar PDF
    elements = [table]
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="horario_{usuario_id}.pdf"'
    
    return response


# -------- Endpoint para exportar Excel --------
@csrf_exempt
def exportar_horario_excel(request):
    """Genera y descarga el horario en formato Excel"""
    if request.method != 'GET':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    usuario_id = request.GET.get('usuario_id')
    if not usuario_id:
        return JsonResponse({"error": "usuario_id es requerido"}, status=400)
    
    horarios, rol_nombre = obtener_horarios_usuario(usuario_id)
    
    if not horarios:
        return JsonResponse({"error": "No se encontraron horarios"}, status=404)
    
    # Generar tabla
    dias, franjas, tabla = generar_tabla_horario(horarios, rol_nombre)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mi Horario"
    
    # Estilos
    header_fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    hour_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    hour_font = Font(bold=True, size=9)
    class_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Escribir header
    ws['A1'] = 'Hora'
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment
    ws['A1'].border = border
    
    for idx, dia in enumerate(dias, start=2):
        cell = ws.cell(row=1, column=idx)
        cell.value = dia
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Escribir datos
    for franja_idx, franja in enumerate(franjas, start=2):
        # Columna de horas
        cell = ws.cell(row=franja_idx, column=1)
        cell.value = franja
        cell.fill = hour_fill
        cell.font = hour_font
        cell.alignment = center_alignment
        cell.border = border
        
        # Celdas de cada día
        for dia_idx in range(len(dias)):
            col = dia_idx + 2
            cell = ws.cell(row=franja_idx, column=col)
            key = (dia_idx, franja_idx - 2)
            cell_text = tabla.get(key, '')
            cell.value = cell_text
            
            if cell_text:
                cell.fill = class_fill
            
            cell.alignment = center_alignment
            cell.border = border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 18
    for col in 'BCDEFG':
        ws.column_dimensions[col].width = 22
    
    # Ajustar alto de filas
    for row in range(2, len(franjas) + 2):
        ws.row_dimensions[row].height = 60
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Preparar respuesta
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="horario_{usuario_id}.xlsx"'
    
    return response


# -------- Endpoint para exportar PDF con array de horarios (POST) --------
@csrf_exempt
def exportar_horarios_pdf_post(request):
    """Genera y descarga horarios en formato PDF desde un array de horarios"""
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        data = json.loads(request.body)
        horarios_data = data.get('horarios', [])
        
        if not horarios_data:
            return JsonResponse({"error": "No se proporcionaron horarios"}, status=400)
        
        # Filtrar solo horarios aprobados
        horarios_data = [h for h in horarios_data if h.get('estado', '').lower() == 'aprobado']
        
        if not horarios_data:
            return JsonResponse({"error": "No hay horarios aprobados para exportar"}, status=400)
        
        # Agrupar horarios por grupo
        grupos_dict = {}
        for h in horarios_data:
            grupo_key = f"{h.get('grupo_nombre', 'N/A')} - {h.get('programa_nombre', 'N/A')}"
            if grupo_key not in grupos_dict:
                grupos_dict[grupo_key] = []
            grupos_dict[grupo_key].append(h)
        
        # Crear PDF con múltiples tablas (una por grupo)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               rightMargin=30, leftMargin=30,
                               topMargin=30, bottomMargin=30)
        
        elements = []
        
        for grupo_key, horarios_grupo in grupos_dict.items():
            # Crear tabla para este grupo con TODAS las horas (6:00 a 21:00)
            dias_dict = {}
            
            # Inicializar estructura de días (mapeo de nombres en español)
            dias_orden = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
            dias_map = {
                'lunes': 'lunes',
                'martes': 'martes',
                'miercoles': 'miercoles',
                'miércoles': 'miercoles',
                'jueves': 'jueves',
                'viernes': 'viernes'
            }
            
            for dia in dias_orden:
                dias_dict[dia] = {}
            
            # Llenar con los horarios disponibles
            for h in horarios_grupo:
                dia_raw = h.get('dia_semana', '').lower().strip()
                hora_inicio_raw = str(h.get('hora_inicio', '')).strip()
                hora_fin_raw = str(h.get('hora_fin', '')).strip()
                
                # Normalizar horas (quitar segundos si los tiene: "10:00:00" -> "10:00")
                def normalizar_hora(hora_str):
                    if ':' in hora_str:
                        partes = hora_str.split(':')
                        return f"{partes[0]}:{partes[1]}"
                    return hora_str
                
                hora_inicio = normalizar_hora(hora_inicio_raw)
                hora_fin = normalizar_hora(hora_fin_raw)
                
                # Normalizar nombre del día
                dia = dias_map.get(dia_raw, dia_raw)
                
                
                if dia and hora_inicio and dia in dias_dict:
                    # Crear contenido de la celda - formato especificado
                    asignatura = str(h.get('asignatura_nombre', '')).strip() or ''
                    docente = str(h.get('docente_nombre', '')).strip() or ''
                    espacio = str(h.get('espacio_nombre', '')).strip() or ''
                    
                    # Formatear según especificación:
                    # Asignatura (Mayúscula inicial) / DOCENTE (MAYÚSCULAS) - Salón
                    if asignatura:
                        # Capitalizar asignatura correctamente (incluyendo números romanos)
                        # Ejemplo: "calculo i" -> "Cálculo I"
                        asignatura_fmt = asignatura.strip()
                        palabras = asignatura_fmt.split()
                        palabras_fmt = []
                        for palabra in palabras:
                            # Si la palabra es un número romano, convertir a mayúsculas
                            if palabra.lower() in ['i', 'ii', 'iii', 'iv', 'v', 'vi', 'vii', 'viii', 'ix', 'x']:
                                palabras_fmt.append(palabra.upper())
                            else:
                                palabras_fmt.append(palabra.capitalize())
                        asignatura_fmt = ' '.join(palabras_fmt)
                        
                        # Docente en mayúsculas
                        docente_fmt = docente.upper() if docente and docente != "Sin asignar" else ""
                        
                        # Salón tal cual
                        espacio_fmt = espacio
                        
                        # Formato con saltos de línea para que no se cruce entre celdas
                        # Asignatura en primera línea
                        # Docente en segunda línea
                        # Salón en tercera línea
                        cell_content = f"{asignatura_fmt}"
                        if docente_fmt:
                            cell_content += f" / {docente_fmt}"
                        if espacio_fmt:
                            cell_content += f"\n{espacio_fmt}"
                        
                        # Agregar la asignatura en TODAS las horas que ocupa
                        try:
                            hora_inicio_h = int(hora_inicio.split(':')[0])
                            hora_fin_h = int(hora_fin.split(':')[0])
                            
                            # Llenar todas las horas desde inicio hasta fin
                            for h_actual in range(hora_inicio_h, hora_fin_h):
                                hora_key = f"{h_actual:02d}:00"
                                dias_dict[dia][hora_key] = cell_content
                            
                        except Exception as e:
                            dias_dict[dia][hora_inicio] = cell_content
            
            # Generar intervalos de horas (6:00-7:00, 7:00-8:00, etc.)
            horas_intervalos = []
            for h in range(6, 21):
                intervalo = f"{h:02d}:00-{h+1:02d}:00"
                horas_intervalos.append((intervalo, f"{h:02d}:00"))  # (etiqueta, clave_busqueda)
            
            # Preparar datos para tabla
            dias_nombres = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
            table_data = [['Hora'] + dias_nombres]
            
            for intervalo, hora_key in horas_intervalos:
                row = [intervalo]
                for dia in dias_orden:
                    cell_text = dias_dict[dia].get(hora_key, '')
                    row.append(cell_text)
                table_data.append(row)
            
            # Crear tabla optimizada para una página
            # Ancho: 1 col hora (0.8") + 5 cols días (2.0" c/u) = 10.8"
            col_widths = [0.8 * inch] + [2.0 * inch] * 5
            # Alto: ajustar según cantidad de filas
            num_filas = len(table_data)
            # Primera fila (header) más compacta, resto para 3 líneas de contenido
            row_heights = [0.35 * inch] + [0.4 * inch] * (num_filas - 1)
            table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
            
            # Estilo optimizado para una página
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#991B1B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F3F4F6')),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (0, -1), 5),
                ('FONTSIZE', (1, 1), (-1, -1), 5),  
                ('TOPPADDING', (1, 1), (-1, -1), 2),
                ('BOTTOMPADDING', (1, 1), (-1, -1), 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ])
            
            # Colorear celdas con clases
            for row_idx, (intervalo, hora_key) in enumerate(horas_intervalos, start=1):
                for col_idx, dia in enumerate(dias_orden, start=1):
                    if hora_key in dias_dict[dia]:
                        style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#DBEAFE'))
            
            table.setStyle(style)
            
            # Agregar título y tabla a elementos
            styles = getSampleStyleSheet()
            title_style = styles['Heading3']
            
            elements.append(Paragraph(f"<b>Grupo: {grupo_key}</b>", title_style))
            elements.append(Spacer(1, 0.1 * inch))
            elements.append(table)
            elements.append(PageBreak())  # Una página por grupo
        
        # Generar PDF
        doc.build(elements)
        
        # Respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="horarios_centro.pdf"'
        
        return response
    
    except Exception as e:
        import traceback
        print(f"ERROR en PDF: {str(e)}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return JsonResponse({"error": str(e)}, status=500)


# -------- Endpoint para exportar Excel con array de horarios (POST) --------
@csrf_exempt
def exportar_horarios_excel_post(request):
    """Genera y descarga horarios en formato Excel desde un array de horarios"""
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        data = json.loads(request.body)
        horarios_data = data.get('horarios', [])
        
        if not horarios_data:
            return JsonResponse({"error": "No se proporcionaron horarios"}, status=400)
        
        # Filtrar solo horarios aprobados
        horarios_data = [h for h in horarios_data if h.get('estado', '').lower() == 'aprobado']
        
        if not horarios_data:
            return JsonResponse({"error": "No hay horarios aprobados para exportar"}, status=400)
        
        # Agrupar horarios por grupo
        grupos_dict = {}
        for h in horarios_data:
            grupo_key = f"{h.get('grupo_nombre', 'N/A')} - {h.get('programa_nombre', 'N/A')}"
            if grupo_key not in grupos_dict:
                grupos_dict[grupo_key] = []
            grupos_dict[grupo_key].append(h)
        
        
        # Crear workbook con múltiples hojas (una por grupo)
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        class_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for grupo_key, horarios_grupo in grupos_dict.items():
            # Crear tabla para este grupo con TODAS las horas (6:00 a 21:00)
            dias_dict = {}
            
            # Inicializar estructura de días (mapeo de nombres en español)
            dias_orden = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
            dias_map = {
                'lunes': 'lunes',
                'martes': 'martes',
                'miercoles': 'miercoles',
                'miércoles': 'miercoles',
                'jueves': 'jueves',
                'viernes': 'viernes'
            }
            
            for dia in dias_orden:
                dias_dict[dia] = {}
            
            # Llenar con los horarios disponibles
            for h in horarios_grupo:
                dia_raw = h.get('dia_semana', '').lower().strip()
                hora_inicio_raw = str(h.get('hora_inicio', '')).strip()
                hora_fin_raw = str(h.get('hora_fin', '')).strip()
                
                # Normalizar horas (quitar segundos si los tiene: "10:00:00" -> "10:00")
                def normalizar_hora(hora_str):
                    if ':' in hora_str:
                        partes = hora_str.split(':')
                        return f"{partes[0]}:{partes[1]}"
                    return hora_str
                
                hora_inicio = normalizar_hora(hora_inicio_raw)
                hora_fin = normalizar_hora(hora_fin_raw)
                
                # Normalizar nombre del día
                dia = dias_map.get(dia_raw, dia_raw)
                
                
                if dia and hora_inicio and dia in dias_dict:
                    # Crear contenido de la celda - formato especificado
                    asignatura = str(h.get('asignatura_nombre', '')).strip() or ''
                    docente = str(h.get('docente_nombre', '')).strip() or ''
                    espacio = str(h.get('espacio_nombre', '')).strip() or ''
                    
                    # Formatear según especificación:
                    # Asignatura (Mayúscula inicial) / DOCENTE (MAYÚSCULAS) - Salón
                    if asignatura:
                        # Capitalizar asignatura correctamente (incluyendo números romanos)
                        # Ejemplo: "calculo i" -> "Cálculo I"
                        asignatura_fmt = asignatura.strip()
                        palabras = asignatura_fmt.split()
                        palabras_fmt = []
                        for palabra in palabras:
                            # Si la palabra es un número romano, convertir a mayúsculas
                            if palabra.lower() in ['i', 'ii', 'iii', 'iv', 'v', 'vi', 'vii', 'viii', 'ix', 'x']:
                                palabras_fmt.append(palabra.upper())
                            else:
                                palabras_fmt.append(palabra.capitalize())
                        asignatura_fmt = ' '.join(palabras_fmt)
                        
                        # Docente en mayúsculas
                        docente_fmt = docente.upper() if docente and docente != "Sin asignar" else ""
                        
                        # Salón tal cual
                        espacio_fmt = espacio
                        
                        # Formato con saltos de línea para que no se cruce entre celdas
                        # Asignatura / Docente en primera línea
                        # Salón en segunda línea
                        cell_content = f"{asignatura_fmt}"
                        if docente_fmt:
                            cell_content += f" / {docente_fmt}"
                        if espacio_fmt:
                            cell_content += f"\n{espacio_fmt}"
                        
                        # Agregar la asignatura en TODAS las horas que ocupa
                        try:
                            hora_inicio_h = int(hora_inicio.split(':')[0])
                            hora_fin_h = int(hora_fin.split(':')[0])
                            
                            # Llenar todas las horas desde inicio hasta fin
                            for h_actual in range(hora_inicio_h, hora_fin_h):
                                hora_key = f"{h_actual:02d}:00"
                                dias_dict[dia][hora_key] = cell_content
                            
                        except Exception as e:
                            dias_dict[dia][hora_inicio] = cell_content
            
            # Generar intervalos de horas (6:00-7:00, 7:00-8:00, etc.)
            horas_intervalos = []
            for h in range(6, 21):
                intervalo = f"{h:02d}:00-{h+1:02d}:00"
                horas_intervalos.append((intervalo, f"{h:02d}:00"))  # (etiqueta, clave_busqueda)
            
            # Crear hoja para este grupo
            ws = wb.create_sheet(title=grupo_key[:31])  # Excel limita nombre a 31 caracteres
            
            # Header con días
            dias_nombres = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
            ws.append(['Hora'] + dias_nombres)
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Datos - intervalos de horas
            for intervalo, hora_key in horas_intervalos:
                row = [intervalo]
                for dia in dias_orden:
                    cell_text = dias_dict[dia].get(hora_key, '')
                    row.append(cell_text)
                ws.append(row)
            
            # Aplicar estilos a datos
            for row_idx, (intervalo, hora_key) in enumerate(horas_intervalos, start=2):
                for col_idx, dia in enumerate(dias_orden, start=2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell_text = dias_dict[dia].get(hora_key, '')
                    
                    cell.value = cell_text
                    
                    if cell_text:
                        cell.fill = class_fill
                    
                    cell.alignment = center_alignment
                    cell.border = border
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 14
            for col in 'BCDEF':
                ws.column_dimensions[col].width = 25
            
            # Ajustar alto de filas para 3 líneas de contenido
            for row in range(2, len(horas_intervalos) + 2):
                ws.row_dimensions[row].height = 75
        
        # Validar que hay al menos una hoja
        if len(wb.sheetnames) == 0:
            print(f"ADVERTENCIA: No hay hojas en el workbook, creando hoja vacía", file=sys.stderr)
            ws = wb.create_sheet(title="Vacío")
            ws.append(['No hay horarios para mostrar'])
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Respuesta
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="horarios_centro.xlsx"'
        
        return response
    
    except Exception as e:
        import traceback
        print(f"ERROR en Excel: {str(e)}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return JsonResponse({"error": str(e)}, status=500)


# -------- Endpoints para descargar horario individual (docente/estudiante) --------
@csrf_exempt
def exportar_pdf_usuario(request):
    """Descarga el horario en PDF para un usuario específico (docente o estudiante)"""
    if request.method != 'GET':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        usuario_id = request.GET.get('usuario_id')
        if not usuario_id:
            return JsonResponse({"error": "usuario_id requerido"}, status=400)
        
        try:
            # Obtener horarios del usuario
            horarios_data = list_horarios_extendidos_usuario(usuario_id)
            
            if not horarios_data:
                return JsonResponse({"error": "No hay horarios registrados"}, status=400)
            
            # Generar PDF directamente
            return generar_pdf_horarios(horarios_data)
        except Exception as inner_e:
            import traceback
            print(f"ERROR en exportar_pdf_usuario (inner): {str(inner_e)}", file=sys.stderr)
            print(traceback.format_exc(), file=sys.stderr)
            return JsonResponse({"error": "No hay horarios registrados"}, status=400)
    
    except Exception as e:
        import traceback
        print(f"ERROR en exportar_pdf_usuario: {str(e)}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def exportar_excel_usuario(request):
    """Descarga el horario en Excel para un usuario específico (docente o estudiante)"""
    if request.method != 'GET':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        usuario_id = request.GET.get('usuario_id')
        if not usuario_id:
            return JsonResponse({"error": "usuario_id requerido"}, status=400)
        
        try:
            # Obtener horarios del usuario
            horarios_data = list_horarios_extendidos_usuario(usuario_id)
            
            if not horarios_data:
                return JsonResponse({"error": "No hay horarios registrados"}, status=400)
            
            # Generar Excel directamente
            return generar_excel_horarios(horarios_data)
        except Exception as inner_e:
            import traceback
            print(f"ERROR en exportar_excel_usuario (inner): {str(inner_e)}", file=sys.stderr)
            print(traceback.format_exc(), file=sys.stderr)
            return JsonResponse({"error": "No hay horarios registrados"}, status=400)
    
    except Exception as e:
        import traceback
        print(f"ERROR en exportar_excel_usuario: {str(e)}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return JsonResponse({"error": str(e)}, status=500)


def generar_pdf_horarios(horarios_data):
    """Genera un PDF con los horarios proporcionados"""
    try:
        # Agrupar horarios por grupo
        grupos_dict = {}
        for h in horarios_data:
            grupo_key = f"{h.get('grupo_nombre', 'N/A')} - {h.get('programa_nombre', 'N/A')}"
            if grupo_key not in grupos_dict:
                grupos_dict[grupo_key] = []
            grupos_dict[grupo_key].append(h)
        
        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               rightMargin=30, leftMargin=30,
                               topMargin=30, bottomMargin=30)
        
        elements = []
        
        for grupo_key, horarios_grupo in grupos_dict.items():
            dias_dict = {}
            dias_orden = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
            dias_map = {
                'lunes': 'lunes',
                'martes': 'martes',
                'miercoles': 'miercoles',
                'miércoles': 'miercoles',
                'jueves': 'jueves',
                'viernes': 'viernes'
            }
            
            for dia in dias_orden:
                dias_dict[dia] = {}
            
            # Llenar con los horarios disponibles
            for h in horarios_grupo:
                dia_raw = h.get('dia_semana', '').lower().strip()
                hora_inicio_raw = str(h.get('hora_inicio', '')).strip()
                hora_fin_raw = str(h.get('hora_fin', '')).strip()
                
                def normalizar_hora(hora_str):
                    if ':' in hora_str:
                        partes = hora_str.split(':')
                        return f"{partes[0]}:{partes[1]}"
                    return hora_str
                
                hora_inicio = normalizar_hora(hora_inicio_raw)
                hora_fin = normalizar_hora(hora_fin_raw)
                dia = dias_map.get(dia_raw, dia_raw)
                
                if dia and hora_inicio and dia in dias_dict:
                    asignatura = str(h.get('asignatura_nombre', '')).strip() or ''
                    docente = str(h.get('docente_nombre', '')).strip() or ''
                    espacio = str(h.get('espacio_nombre', '')).strip() or ''
                    
                    if asignatura:
                        asignatura_fmt = asignatura.strip()
                        palabras = asignatura_fmt.split()
                        palabras_fmt = []
                        for palabra in palabras:
                            if palabra.lower() in ['i', 'ii', 'iii', 'iv', 'v', 'vi', 'vii', 'viii', 'ix', 'x']:
                                palabras_fmt.append(palabra.upper())
                            else:
                                palabras_fmt.append(palabra.capitalize())
                        asignatura_fmt = ' '.join(palabras_fmt)
                        
                        docente_fmt = docente.upper() if docente and docente != "Sin asignar" else ""
                        espacio_fmt = espacio
                        
                        cell_content = f"{asignatura_fmt}"
                        if docente_fmt:
                            cell_content += f" / {docente_fmt}"
                        if espacio_fmt:
                            cell_content += f"\n{espacio_fmt}"
                        
                        try:
                            hora_inicio_h = int(hora_inicio.split(':')[0])
                            hora_fin_h = int(hora_fin.split(':')[0])
                            
                            for h_actual in range(hora_inicio_h, hora_fin_h):
                                hora_key = f"{h_actual:02d}:00"
                                dias_dict[dia][hora_key] = cell_content
                        except Exception as e:
                            dias_dict[dia][hora_inicio] = cell_content
            
            # Generar intervalos de horas
            horas_intervalos = []
            for h in range(6, 21):
                intervalo = f"{h:02d}:00-{h+1:02d}:00"
                horas_intervalos.append((intervalo, f"{h:02d}:00"))
            
            # Preparar datos para tabla
            dias_nombres = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
            table_data = [['Hora'] + dias_nombres]
            
            for intervalo, hora_key in horas_intervalos:
                row = [intervalo]
                for dia in dias_orden:
                    cell_text = dias_dict[dia].get(hora_key, '')
                    row.append(cell_text)
                table_data.append(row)
            
            # Crear tabla
            col_widths = [0.8 * inch] + [2.0 * inch] * 5
            num_filas = len(table_data)
            row_heights = [0.35 * inch] + [0.4 * inch] * (num_filas - 1)
            table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
            
            # Estilo
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#991B1B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F3F4F6')),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (0, -1), 5),
                ('FONTSIZE', (1, 1), (-1, -1), 5),
                ('TOPPADDING', (1, 1), (-1, -1), 2),
                ('BOTTOMPADDING', (1, 1), (-1, -1), 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ])
            
            for row_idx, (intervalo, hora_key) in enumerate(horas_intervalos, start=1):
                for col_idx, dia in enumerate(dias_orden, start=1):
                    if hora_key in dias_dict[dia]:
                        style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#DBEAFE'))
            
            table.setStyle(style)
            
            styles = getSampleStyleSheet()
            title_style = styles['Heading3']
            
            elements.append(Paragraph(f"<b>Grupo: {grupo_key}</b>", title_style))
            elements.append(Spacer(1, 0.1 * inch))
            elements.append(table)
            elements.append(PageBreak())
        
        # Generar PDF
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="horarios.pdf"'
        return response
    
    except Exception as e:
        import traceback
        print(f"ERROR en generar_pdf_horarios: {str(e)}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return JsonResponse({"error": str(e)}, status=500)


def generar_excel_horarios(horarios_data):
    """Genera un Excel con los horarios proporcionados"""
    try:
        # Agrupar horarios por grupo
        grupos_dict = {}
        for h in horarios_data:
            grupo_key = f"{h.get('grupo_nombre', 'N/A')} - {h.get('programa_nombre', 'N/A')}"
            if grupo_key not in grupos_dict:
                grupos_dict[grupo_key] = []
            grupos_dict[grupo_key].append(h)
        
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        class_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for grupo_key, horarios_grupo in grupos_dict.items():
            dias_dict = {}
            dias_orden = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
            dias_map = {
                'lunes': 'lunes',
                'martes': 'martes',
                'miercoles': 'miercoles',
                'miércoles': 'miercoles',
                'jueves': 'jueves',
                'viernes': 'viernes'
            }
            
            for dia in dias_orden:
                dias_dict[dia] = {}
            
            # Llenar con los horarios disponibles
            for h in horarios_grupo:
                dia_raw = h.get('dia_semana', '').lower().strip()
                hora_inicio_raw = str(h.get('hora_inicio', '')).strip()
                hora_fin_raw = str(h.get('hora_fin', '')).strip()
                
                def normalizar_hora(hora_str):
                    if ':' in hora_str:
                        partes = hora_str.split(':')
                        return f"{partes[0]}:{partes[1]}"
                    return hora_str
                
                hora_inicio = normalizar_hora(hora_inicio_raw)
                hora_fin = normalizar_hora(hora_fin_raw)
                dia = dias_map.get(dia_raw, dia_raw)
                
                if dia and hora_inicio and dia in dias_dict:
                    asignatura = str(h.get('asignatura_nombre', '')).strip() or ''
                    docente = str(h.get('docente_nombre', '')).strip() or ''
                    espacio = str(h.get('espacio_nombre', '')).strip() or ''
                    
                    if asignatura:
                        asignatura_fmt = asignatura.strip()
                        palabras = asignatura_fmt.split()
                        palabras_fmt = []
                        for palabra in palabras:
                            if palabra.lower() in ['i', 'ii', 'iii', 'iv', 'v', 'vi', 'vii', 'viii', 'ix', 'x']:
                                palabras_fmt.append(palabra.upper())
                            else:
                                palabras_fmt.append(palabra.capitalize())
                        asignatura_fmt = ' '.join(palabras_fmt)
                        
                        docente_fmt = docente.upper() if docente and docente != "Sin asignar" else ""
                        espacio_fmt = espacio
                        
                        cell_content = f"{asignatura_fmt}"
                        if docente_fmt:
                            cell_content += f" / {docente_fmt}"
                        if espacio_fmt:
                            cell_content += f"\n{espacio_fmt}"
                        
                        try:
                            hora_inicio_h = int(hora_inicio.split(':')[0])
                            hora_fin_h = int(hora_fin.split(':')[0])
                            
                            for h_actual in range(hora_inicio_h, hora_fin_h):
                                hora_key = f"{h_actual:02d}:00"
                                dias_dict[dia][hora_key] = cell_content
                        except Exception as e:
                            dias_dict[dia][hora_inicio] = cell_content
            
            # Generar intervalos de horas
            horas_intervalos = []
            for h in range(6, 21):
                intervalo = f"{h:02d}:00-{h+1:02d}:00"
                horas_intervalos.append((intervalo, f"{h:02d}:00"))
            
            # Crear hoja
            ws = wb.create_sheet(title=grupo_key[:31])
            
            # Header
            dias_nombres = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
            ws.append(['Hora'] + dias_nombres)
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Datos
            for intervalo, hora_key in horas_intervalos:
                row = [intervalo]
                for dia in dias_orden:
                    cell_text = dias_dict[dia].get(hora_key, '')
                    row.append(cell_text)
                ws.append(row)
            
            # Aplicar estilos
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = border
                    if cell.value:
                        cell.fill = class_fill
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 14
            for col in 'BCDEF':
                ws.column_dimensions[col].width = 25
            
            # Ajustar alto de filas
            for row in range(2, len(horas_intervalos) + 2):
                ws.row_dimensions[row].height = 75
        
        # Validar que hay al menos una hoja
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="Vacío")
            ws.append(['No hay horarios para mostrar'])
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Respuesta
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="horarios.xlsx"'
        
        return response
    
    except Exception as e:
        import traceback
        print(f"ERROR en generar_excel_horarios: {str(e)}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return JsonResponse({"error": str(e)}, status=500)


def list_horarios_extendidos_usuario(usuario_id):
    """Obtiene los horarios extendidos para un usuario específico (docente o estudiante)"""
    try:
        from django.contrib.auth.models import User
        import traceback
        
        usuario = User.objects.get(id=usuario_id)
        
        # Obtener el perfil del usuario para determinar si es docente o estudiante
        from usuarios.models import Perfil
        perfil = Perfil.objects.get(usuario=usuario)
        
        horarios = []
        
        if perfil.rol.nombre == 'docente':
            # Si es docente, obtener sus horarios
            from docentes.models import Docente
            try:
                docente = Docente.objects.get(usuario=usuario)
                items = Horario.objects.filter(docente=docente, estado='aprobado').select_related(
                    'grupo', 'asignatura', 'docente', 'espacio', 'grupo__programa'
                )
            except Docente.DoesNotExist:
                print(f"ADVERTENCIA: No se encontró Docente para usuario {usuario_id}", file=sys.stderr)
                items = []
        else:
            # Si es estudiante, obtener horarios de su grupo
            from estudiantes.models import Estudiante
            try:
                estudiante = Estudiante.objects.get(usuario=usuario)
                items = Horario.objects.filter(grupo=estudiante.grupo, estado='aprobado').select_related(
                    'grupo', 'asignatura', 'docente', 'espacio', 'grupo__programa'
                )
            except Estudiante.DoesNotExist:
                print(f"ADVERTENCIA: No se encontró Estudiante para usuario {usuario_id}", file=sys.stderr)
                items = []
        
        for i in items:
            horarios.append({
                "id": i.id,
                "grupo_id": i.grupo.id,
                "grupo_nombre": i.grupo.nombre,
                "programa_id": i.grupo.programa.id,
                "programa_nombre": i.grupo.programa.nombre,
                "semestre": i.grupo.semestre,
                "asignatura_id": i.asignatura.id,
                "asignatura_nombre": i.asignatura.nombre,
                "docente_id": (i.docente.id if i.docente else None),
                "docente_nombre": i.docente.nombre if i.docente else "Sin asignar",
                "espacio_id": i.espacio.id,
                "espacio_nombre": i.espacio.nombre,
                "dia_semana": i.dia_semana,
                "hora_inicio": str(i.hora_inicio),
                "hora_fin": str(i.hora_fin),
                "cantidad_estudiantes": i.cantidad_estudiantes
            })
        
        return horarios
    
    except Exception as e:
        import traceback
        print(f"ERROR en list_horarios_extendidos_usuario: {str(e)}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return []


# -------- Endpoints para exportar PDF y Excel por Docente --------
@csrf_exempt
def exportar_horarios_pdf_docente(request):
    """Genera y descarga horarios en formato PDF agrupados por docente"""
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        data = json.loads(request.body)
        horarios_data = data.get('horarios', [])
        
        if not horarios_data:
            return JsonResponse({"error": "No se proporcionaron horarios"}, status=400)
        
        # Filtrar solo horarios aprobados
        horarios_data = [h for h in horarios_data if h.get('estado', '').lower() == 'aprobado']
        
        if not horarios_data:
            return JsonResponse({"error": "No hay horarios aprobados para exportar"}, status=400)
        
        # Agrupar horarios por docente
        docentes_dict = {}
        for h in horarios_data:
            docente_key = h.get('docente_nombre', 'Sin asignar')
            facultad = h.get('programa_nombre', 'N/A')
            
            if docente_key not in docentes_dict:
                docentes_dict[docente_key] = {
                    'facultad': facultad,
                    'horarios': []
                }
            docentes_dict[docente_key]['horarios'].append(h)
        
        # Crear PDF con múltiples tablas (una por docente)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               rightMargin=30, leftMargin=30,
                               topMargin=30, bottomMargin=30)
        
        elements = []
        
        for docente_nombre, docente_info in docentes_dict.items():
            facultad = docente_info['facultad']
            horarios_docente = docente_info['horarios']
            
            # Crear tabla para este docente
            dias_dict = {}
            
            # Inicializar estructura de días
            dias_orden = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
            dias_map = {
                'lunes': 'lunes',
                'martes': 'martes',
                'miercoles': 'miercoles',
                'miércoles': 'miercoles',
                'jueves': 'jueves',
                'viernes': 'viernes'
            }
            
            for dia in dias_orden:
                dias_dict[dia] = {}
            
            # Llenar con los horarios disponibles
            for h in horarios_docente:
                dia_raw = h.get('dia_semana', '').lower().strip()
                hora_inicio_raw = str(h.get('hora_inicio', '')).strip()
                hora_fin_raw = str(h.get('hora_fin', '')).strip()
                
                # Normalizar horas
                def normalizar_hora(hora_str):
                    if ':' in hora_str:
                        partes = hora_str.split(':')
                        return f"{partes[0]}:{partes[1]}"
                    return hora_str
                
                hora_inicio = normalizar_hora(hora_inicio_raw)
                hora_fin = normalizar_hora(hora_fin_raw)
                
                # Normalizar nombre del día
                dia = dias_map.get(dia_raw, dia_raw)
                
                if dia and hora_inicio and dia in dias_dict:
                    # Crear contenido de la celda
                    asignatura = str(h.get('asignatura_nombre', '')).strip() or ''
                    grupo = str(h.get('grupo_nombre', '')).strip() or ''
                    espacio = str(h.get('espacio_nombre', '')).strip() or ''
                    
                    if asignatura:
                        # Capitalizar asignatura
                        asignatura_fmt = asignatura.strip()
                        palabras = asignatura_fmt.split()
                        palabras_fmt = []
                        for palabra in palabras:
                            if palabra.lower() in ['i', 'ii', 'iii', 'iv', 'v', 'vi', 'vii', 'viii', 'ix', 'x']:
                                palabras_fmt.append(palabra.upper())
                            else:
                                palabras_fmt.append(palabra.capitalize())
                        asignatura_fmt = ' '.join(palabras_fmt)
                        
                        # Formato: Asignatura / GRUPO - Salón
                        cell_content = f"{asignatura_fmt}"
                        if grupo:
                            cell_content += f" / {grupo}"
                        if espacio:
                            cell_content += f"\n{espacio}"
                        
                        # Agregar la asignatura en TODAS las horas que ocupa
                        try:
                            hora_inicio_h = int(hora_inicio.split(':')[0])
                            hora_fin_h = int(hora_fin.split(':')[0])
                            
                            for h_actual in range(hora_inicio_h, hora_fin_h):
                                hora_key = f"{h_actual:02d}:00"
                                dias_dict[dia][hora_key] = cell_content
                            
                        except Exception as e:
                            dias_dict[dia][hora_inicio] = cell_content
            
            # Generar intervalos de horas
            horas_intervalos = []
            for h in range(6, 21):
                intervalo = f"{h:02d}:00-{h+1:02d}:00"
                horas_intervalos.append((intervalo, f"{h:02d}:00"))
            
            # Preparar datos para tabla
            dias_nombres = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
            table_data = [['Hora'] + dias_nombres]
            
            for intervalo, hora_key in horas_intervalos:
                row = [intervalo]
                for dia in dias_orden:
                    cell_text = dias_dict[dia].get(hora_key, '')
                    row.append(cell_text)
                table_data.append(row)
            
            # Crear tabla
            col_widths = [0.8 * inch] + [2.0 * inch] * 5
            num_filas = len(table_data)
            row_heights = [0.35 * inch] + [0.4 * inch] * (num_filas - 1)
            table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
            
            # Estilo
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#991B1B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F3F4F6')),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (0, -1), 5),
                ('FONTSIZE', (1, 1), (-1, -1), 5),
                ('TOPPADDING', (1, 1), (-1, -1), 2),
                ('BOTTOMPADDING', (1, 1), (-1, -1), 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ])
            
            # Colorear celdas con clases
            for row_idx, (intervalo, hora_key) in enumerate(horas_intervalos, start=1):
                for col_idx, dia in enumerate(dias_orden, start=1):
                    if hora_key in dias_dict[dia]:
                        style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#DBEAFE'))
            
            table.setStyle(style)
            
            # Agregar título y tabla a elementos
            styles = getSampleStyleSheet()
            title_style = styles['Heading3']
            
            elements.append(Paragraph(f"<b>Docente: {docente_nombre}</b>", title_style))
            elements.append(Paragraph(f"<b>Facultad: {facultad}</b>", title_style))
            elements.append(Spacer(1, 0.1 * inch))
            elements.append(table)
            elements.append(PageBreak())
        
        # Generar PDF
        doc.build(elements)
        
        # Respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="horarios_docentes.pdf"'
        
        return response
    
    except Exception as e:
        import traceback
        print(f"ERROR en exportar_horarios_pdf_docente: {str(e)}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def exportar_horarios_excel_docente(request):
    """Genera y descarga horarios en formato Excel agrupados por docente"""
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        data = json.loads(request.body)
        horarios_data = data.get('horarios', [])
        
        if not horarios_data:
            return JsonResponse({"error": "No se proporcionaron horarios"}, status=400)
        
        # Filtrar solo horarios aprobados
        horarios_data = [h for h in horarios_data if h.get('estado', '').lower() == 'aprobado']
        
        if not horarios_data:
            return JsonResponse({"error": "No hay horarios aprobados para exportar"}, status=400)
        
        # Agrupar horarios por docente
        docentes_dict = {}
        for h in horarios_data:
            docente_key = h.get('docente_nombre', 'Sin asignar')
            facultad = h.get('programa_nombre', 'N/A')
            
            if docente_key not in docentes_dict:
                docentes_dict[docente_key] = {
                    'facultad': facultad,
                    'horarios': []
                }
            docentes_dict[docente_key]['horarios'].append(h)
        
        # Crear Excel con múltiples hojas (una por docente)
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remover hoja por defecto
        
        for docente_nombre, docente_info in docentes_dict.items():
            facultad = docente_info['facultad']
            horarios_docente = docente_info['horarios']
            
            # Crear tabla para este docente
            dias_dict = {}
            
            # Inicializar estructura de días
            dias_orden = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
            dias_map = {
                'lunes': 'lunes',
                'martes': 'martes',
                'miercoles': 'miercoles',
                'miércoles': 'miercoles',
                'jueves': 'jueves',
                'viernes': 'viernes'
            }
            
            for dia in dias_orden:
                dias_dict[dia] = {}
            
            # Llenar con los horarios disponibles
            for h in horarios_docente:
                dia_raw = h.get('dia_semana', '').lower().strip()
                hora_inicio_raw = str(h.get('hora_inicio', '')).strip()
                hora_fin_raw = str(h.get('hora_fin', '')).strip()
                
                # Normalizar horas
                def normalizar_hora(hora_str):
                    if ':' in hora_str:
                        partes = hora_str.split(':')
                        return f"{partes[0]}:{partes[1]}"
                    return hora_str
                
                hora_inicio = normalizar_hora(hora_inicio_raw)
                hora_fin = normalizar_hora(hora_fin_raw)
                
                # Normalizar nombre del día
                dia = dias_map.get(dia_raw, dia_raw)
                
                if dia and hora_inicio and dia in dias_dict:
                    # Crear contenido de la celda
                    asignatura = str(h.get('asignatura_nombre', '')).strip() or ''
                    grupo = str(h.get('grupo_nombre', '')).strip() or ''
                    espacio = str(h.get('espacio_nombre', '')).strip() or ''
                    
                    if asignatura:
                        # Capitalizar asignatura
                        asignatura_fmt = asignatura.strip()
                        palabras = asignatura_fmt.split()
                        palabras_fmt = []
                        for palabra in palabras:
                            if palabra.lower() in ['i', 'ii', 'iii', 'iv', 'v', 'vi', 'vii', 'viii', 'ix', 'x']:
                                palabras_fmt.append(palabra.upper())
                            else:
                                palabras_fmt.append(palabra.capitalize())
                        asignatura_fmt = ' '.join(palabras_fmt)
                        
                        # Formato: Asignatura / GRUPO - Salón
                        cell_content = f"{asignatura_fmt}"
                        if grupo:
                            cell_content += f" / {grupo}"
                        if espacio:
                            cell_content += f" - {espacio}"
                        
                        # Agregar la asignatura en TODAS las horas que ocupa
                        try:
                            hora_inicio_h = int(hora_inicio.split(':')[0])
                            hora_fin_h = int(hora_fin.split(':')[0])
                            
                            for h_actual in range(hora_inicio_h, hora_fin_h):
                                hora_key = f"{h_actual:02d}:00"
                                dias_dict[dia][hora_key] = cell_content
                            
                        except Exception as e:
                            dias_dict[dia][hora_inicio] = cell_content
            
            # Generar intervalos de horas
            horas_intervalos = []
            for h in range(6, 21):
                intervalo = f"{h:02d}:00-{h+1:02d}:00"
                horas_intervalos.append((intervalo, f"{h:02d}:00"))
            
            # Preparar datos para tabla
            dias_nombres = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
            table_data = [['Hora'] + dias_nombres]
            
            for intervalo, hora_key in horas_intervalos:
                row = [intervalo]
                for dia in dias_orden:
                    cell_text = dias_dict[dia].get(hora_key, '')
                    row.append(cell_text)
                table_data.append(row)
            
            # Crear hoja en Excel
            sheet_name = f"{docente_nombre[:30]}"  # Limitar nombre de hoja a 30 caracteres
            ws = workbook.create_sheet(title=sheet_name)
            
            # Definir estilos (igual que en Centro de Horarios)
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Rojo oscuro
            class_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # Azul claro
            hora_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # Gris claro
            title_font = Font(bold=True, size=14, color="000000")  # Negro
            hora_font = Font(bold=True, size=10, color="333333")  # Gris oscuro
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # Agregar título con docente y facultad
            ws['A1'] = f"Docente: {docente_nombre}"
            ws['A1'].font = title_font
            ws['A1'].alignment = left_alignment
            
            ws['A2'] = f"Facultad: {facultad}"
            ws['A2'].font = title_font
            ws['A2'].alignment = left_alignment
            
            ws['A3'] = ""  # Espacio en blanco
            
            # Agregar tabla a partir de fila 4
            for row_idx, row_data in enumerate(table_data, start=4):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    # Aplicar estilos al header (primera fila de datos - fila 4)
                    if row_idx == 4:
                        # Todo el header en rojo
                        cell.fill = header_fill
                        cell.font = header_font
                    # Columna de hora (col_idx == 1) en datos - gris claro
                    elif col_idx == 1:
                        cell.fill = hora_fill
                        cell.font = hora_font
                    # Aplicar color azul a las celdas con contenido (clases)
                    elif cell_value and cell_value.strip():
                        cell.fill = class_fill
                        cell.font = Font(bold=False, size=10)
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 14
            for col_idx in range(2, 7):
                ws.column_dimensions[get_column_letter(col_idx)].width = 25
            
            # Ajustar altura de filas
            ws.row_dimensions[1].height = 25
            ws.row_dimensions[2].height = 25
            for row_idx in range(4, 4 + len(table_data)):
                ws.row_dimensions[row_idx].height = 75  # Altura mayor para mostrar contenido multilínea
        
        # Guardar Excel
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="horarios_docentes.xlsx"'
        
        return response
    
    except Exception as e:
        import traceback
        print(f"ERROR en exportar_horarios_excel_docente: {str(e)}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return JsonResponse({"error": str(e)}, status=500)

# ========== Endpoints para Solicitudes de Espacio ==========

@csrf_exempt
def list_solicitudes_espacio(request):
    """Lista todas las solicitudes de espacio con filtros opcionales"""
    if request.method != 'GET':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        estado = request.GET.get('estado')  # Filtro opcional: pendiente, aprobada, rechazada
        
        solicitudes = SolicitudEspacio.objects.select_related(
            'grupo', 'asignatura', 'docente', 'espacio_solicitado', 'planificador', 'aprobado_por'
        ).all()
        
        if estado:
            solicitudes = solicitudes.filter(estado=estado)
        
        lst = []
        for s in solicitudes:
            lst.append({
                "id": s.id,
                "grupo_id": s.grupo.id,
                "grupo_nombre": s.grupo.nombre,
                "asignatura_id": s.asignatura.id,
                "asignatura_nombre": s.asignatura.nombre,
                "docente_id": s.docente.id if s.docente else None,
                "docente_nombre": s.docente.nombre if s.docente else None,
                "espacio_solicitado_id": s.espacio_solicitado.id,
                "espacio_solicitado_nombre": s.espacio_solicitado.nombre,
                "planificador_id": s.planificador.id if s.planificador else None,
                "planificador_nombre": s.planificador.nombre if s.planificador else None,
                "dia_semana": s.dia_semana,
                "hora_inicio": str(s.hora_inicio),
                "hora_fin": str(s.hora_fin),
                "cantidad_estudiantes": s.cantidad_estudiantes,
                "estado": s.estado,
                "fecha_solicitud": str(s.fecha_solicitud),
                "fecha_aprobacion": str(s.fecha_aprobacion) if s.fecha_aprobacion else None,
                "aprobado_por_id": s.aprobado_por.id if s.aprobado_por else None,
                "aprobado_por_nombre": s.aprobado_por.nombre if s.aprobado_por else None,
                "comentario": s.comentario,
                "horario_generado_id": s.horario_generado.id if s.horario_generado else None
            })
        
        return JsonResponse({"solicitudes": lst}, status=200)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def aprobar_solicitud_espacio(request):
    """Aprueba una solicitud de espacio y crea el horario"""
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        data = json.loads(request.body)
        solicitud_id = data.get('solicitud_id')
        admin_id = data.get('admin_id')
        comentario = data.get('comentario', '')
        
        if not solicitud_id or not admin_id:
            return JsonResponse({"error": "solicitud_id y admin_id son requeridos"}, status=400)
        
        solicitud = SolicitudEspacio.objects.get(id=solicitud_id)
        admin = Usuario.objects.get(id=admin_id)
        
        # Si ya existe un horario generado, eliminarlo para crear uno nuevo
        if solicitud.horario_generado:
            solicitud.horario_generado.delete()
            solicitud.horario_generado = None
        
        # Validar conflictos de horarios
        # Verificar si hay otro horario en el mismo espacio, día y hora
        conflictos = Horario.objects.filter(
            espacio=solicitud.espacio_solicitado,
            dia_semana=solicitud.dia_semana,
            estado='aprobado'
        )
        
        for horario_existente in conflictos:
            # Verificar si hay solapamiento de horas
            if not (solicitud.hora_fin <= horario_existente.hora_inicio or 
                    solicitud.hora_inicio >= horario_existente.hora_fin):
                return JsonResponse({
                    "error": f"Conflicto de horario: El espacio {solicitud.espacio_solicitado.nombre} ya está ocupado el {solicitud.dia_semana} de {horario_existente.hora_inicio} a {horario_existente.hora_fin}"
                }, status=409)
        
        # Crear el horario
        horario = Horario(
            grupo=solicitud.grupo,
            asignatura=solicitud.asignatura,
            docente=solicitud.docente,
            espacio=solicitud.espacio_solicitado,
            dia_semana=solicitud.dia_semana,
            hora_inicio=solicitud.hora_inicio,
            hora_fin=solicitud.hora_fin,
            cantidad_estudiantes=solicitud.cantidad_estudiantes,
            estado='aprobado'
        )
        horario.save()
        
        # Actualizar solicitud
        solicitud.estado = 'aprobada'
        solicitud.horario_generado = horario
        solicitud.aprobado_por = admin
        solicitud.fecha_aprobacion = datetime.datetime.now()
        solicitud.comentario = comentario
        solicitud.save()
        
        # Crear notificación para el planificador
        if solicitud.planificador:
            crear_notificacion(
                id_usuario=solicitud.planificador.id,
                tipo='solicitud_aprobada',
                mensaje=f'✅ Tu solicitud de espacio para {solicitud.asignatura.nombre} - Grupo {solicitud.grupo.nombre} ha sido aprobada',
                prioridad='alta'
            )
        
        return JsonResponse({
            "message": "Solicitud aprobada y horario creado",
            "solicitud_id": solicitud.id,
            "horario_id": horario.id
        }, status=200)
    
    except SolicitudEspacio.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada"}, status=404)
    except Usuario.DoesNotExist:
        return JsonResponse({"error": "Admin no encontrado"}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def rechazar_solicitud_espacio(request):
    """Rechaza una solicitud de espacio"""
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    try:
        data = json.loads(request.body)
        solicitud_id = data.get('solicitud_id')
        admin_id = data.get('admin_id')
        comentario = data.get('comentario', '')
        
        if not solicitud_id or not admin_id:
            return JsonResponse({"error": "solicitud_id y admin_id son requeridos"}, status=400)
        
        solicitud = SolicitudEspacio.objects.get(id=solicitud_id)
        admin = Usuario.objects.get(id=admin_id)
        
        # Si ya existe un horario generado, eliminarlo (porque se está rechazando)
        if solicitud.horario_generado:
            solicitud.horario_generado.delete()
            solicitud.horario_generado = None
        
        # Actualizar solicitud
        solicitud.estado = 'rechazada'
        solicitud.aprobado_por = admin
        solicitud.fecha_aprobacion = datetime.datetime.now()
        solicitud.comentario = comentario
        solicitud.save()
        
        # Crear notificación para el planificador
        if solicitud.planificador:
            # Construir mensaje estructurado para mejor legibilidad
            mensaje_rechazo = f'❌ Tu solicitud de espacio para {solicitud.asignatura.nombre} - Grupo {solicitud.grupo.nombre} ha sido rechazada'
            if comentario:
                mensaje_rechazo += f'\n\n📋 Motivo:\n{comentario}'
            crear_notificacion(
                id_usuario=solicitud.planificador.id,
                tipo='solicitud_rechazada',
                mensaje=mensaje_rechazo,
                prioridad='alta'
            )
        
        return JsonResponse({
            "message": "Solicitud rechazada",
            "solicitud_id": solicitud.id
        }, status=200)
    
    except SolicitudEspacio.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada"}, status=404)
    except Usuario.DoesNotExist:
        return JsonResponse({"error": "Admin no encontrado"}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
